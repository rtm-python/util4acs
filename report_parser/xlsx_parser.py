"""
This module implements parsing of person passing data from xlxs files into list of person accessing into restricted area data.

Classes:

"""

from dataclasses import dataclass
from typing import List
from openpyxl import load_workbook
from report_parser import logger
from report_parser import AccessData
from report_parser import EmployeeAccess
from report_parser import Parser
from json import JSONEncoder
from datetime import datetime
from datetime import date
from datetime import time


class DateTimeEncoder(JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()


@dataclass
class XLSXParser(Parser):
    def parse(self, xlsx_filepath: str) -> List[EmployeeAccess]:
        """
        Reads and parses xlsx file, then returns list of access data.

        Args:
            xlsx_filepath   xlsx file path
        """
        logger.info(f"Parsing file: {xlsx_filepath}")
        result: List[EmployeeAccess] = []
        wb = load_workbook(filename=xlsx_filepath)
        for ws in wb:
            for row in ws.iter_rows(
                min_row=self.config["row_first"],
                min_col=1,
                max_row=ws.max_row + 1,
                max_col=ws.max_column + 1,
            ):
                # Read cell values according to configuration
                values = [None] + [
                    cell.value
                    if not isinstance(cell.value, str)
                    else cell.value.strip()
                    for cell in row
                ]
                if len(values) < ws.max_column + 1 or values[1] is None:
                    continue
                (
                    name,
                    unit,
                    e_date,
                    e_time,
                    e_status,
                    e_turnstyle,
                    e_direction,
                    e_area,
                ) = (
                    values[self.config["col_name"]],
                    values[self.config["col_unit"]],
                    values[self.config["col_date"]],
                    values[self.config["col_time"]],
                    values[self.config["col_status"]],
                    values[self.config["col_turnstyle"]],
                    values[self.config["col_direction"]],
                    values[self.config["col_area"]],
                )
                # Add new employee access or find earlier added
                employee_access = None
                for item in result:
                    if item.name == name and item.unit == unit:
                        employee_access = item
                        break
                if employee_access is None:
                    employee_access = EmployeeAccess(name, unit, [])
                    result += [employee_access]
                # Parse date and time from string if needed
                if not isinstance(e_date, date):
                    logger.warn(f"{e_date} ({type(e_date)})")
                    e_date = datetime.strptime(e_date, "%d.%m.%Y")
                if not isinstance(e_time, time):
                    logger.warn(f"{e_time} ({type(e_time)})")
                    e_time = datetime.strptime(e_time, "%H:%M:%S")
                # Create datetime object
                e_datetime = datetime(
                    e_date.year,
                    e_date.month,
                    e_date.day,
                    e_time.hour,
                    e_time.minute,
                    e_time.second,
                )
                # Create access data object within employee access object
                e_status = e_status.upper()
                e_direction = e_direction.upper()
                direction = 0
                if e_status.startswith(self.config["status_success_prefix"]):
                    for direction_prefix in self.config["direction_plus_prefix"]:
                        if e_direction.startswith(direction_prefix):
                            direction = 1
                    for direction_prefix in self.config["direction_minus_prefix"]:
                        if e_direction.startswith(direction_prefix):
                            direction = -1
                if direction == 1:
                    employee_access.access_data_list.append(
                        AccessData(e_area, e_datetime, e_turnstyle, None, None)
                    )
                elif direction == -1:
                    if len(employee_access.access_data_list) == 0:
                        logger.error(
                            f"Exit out before enter in: {name}, {unit}, {e_direction}, {e_datetime}, {e_turnstyle}"
                        )
                        continue
                    access_data = employee_access.access_data_list[-1]
                    if (
                        access_data.exit_out is not None
                        or access_data.exit_out_turnstyle is not None
                    ):
                        continue
                    access_data.exit_out = e_datetime
                    access_data.exit_out_turnstyle = e_turnstyle
        wb.close()
        return result
