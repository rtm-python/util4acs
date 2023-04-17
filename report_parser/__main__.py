"""
This is an additional module to demonstrate package functionality (parses reports from filepaths and creates JSON data file).

Example:
    $ python -m report_parser <reports_folder>
"""

import json
from datetime import date
from pathlib import Path
from typing import List, Dict
from report_parser import logger
from report_parser import Parser
from report_parser import EmployeeAccess
from report_parser.xlsx_parser import XLSXParser
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

with open("report_parser/xlsx_parser.json", "r") as file:
    xlsx_config = json.load(file)

parsers: Dict[str, Parser] = {".xlsx": XLSXParser(xlsx_config)}


def parse(path: Path) -> List[EmployeeAccess]:
    """
    Defines parser by path suffix (filetype) and parses with parser, then returns list of access data.

    Args:
        path    Path object to detect file suffix and absolute path.

    Returns:
        List of access data.
    """
    parser = parsers.get(path.suffix)
    if parser is None:
        return
    result = parser.parse(path.resolve())
    units = []
    for item in result:
        unit = item.unit
        if unit in units:
            continue
        units.append(unit)
    if len(units) > 1:
        logger.error(f"More than one unit in report ({path}) detected: {units}")
    return result


def main(report_filepaths: List[str]) -> None:
    """This function parses reports from defined filepaths and creates JSON file with structured data.

    Args:
        report_filepaths  filepaths where reports from ACS stored

    Returns:
        None
    """
    logger.info(f'Package "{__package__}" demonstration')
    # Parse reports
    employee_access_list: List[EmployeeAccess] = []
    parsed_count = 0
    for filepath in report_filepaths:
        path = Path(filepath)
        if path.is_file():
            access_data_list = parse(path)
            if access_data_list is None:
                continue
            employee_access_list += access_data_list
            parsed_count += 1
        elif path.is_dir():
            for item in path.rglob("*"):
                if item.is_file():
                    access_data_list = parse(item)
                    if access_data_list is None:
                        continue
                    employee_access_list += access_data_list
                    parsed_count += 1
    # Create workbook, pre-define dates and units lists
    wb = Workbook()
    dates = []
    units = []
    for employee_access in employee_access_list:
        if employee_access.unit not in units:
            units.append(employee_access.unit)
        for access_data in employee_access.access_data_list:
            if access_data.exit_out is None:
                continue
            e_date = date(
                access_data.exit_out.year,
                access_data.exit_out.month,
                access_data.exit_out.day,
            )
            if e_date in dates:
                continue
            dates.append(e_date)
    date_row = {}
    for row, e_date in enumerate(sorted(dates), 3):
        date_row[e_date] = row
    unit_ws = {}
    for unit in units:
        ws = wb.create_sheet(unit)
        unit_ws[unit] = ws
        for e_date, row in date_row.items():
            ws.cell(row=row, column=1, value=e_date)
            ws.cell(row=row, column=1).number_format = "dd.mm.yyyy"
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="ffffcc00")
    # Fill unit worksheets with employee access data
    for employee_access in employee_access_list:
        ws = unit_ws[employee_access.unit]
        employee_column = ws.max_column + 1
        ws.cell(
            row=1,
            column=employee_column,
            value=employee_access.name,
        )
        ws.cell(row=1, column=employee_column).fill = PatternFill(
            "solid", fgColor="ffa4ffa4"
        )
        ws.cell(
            row=2,
            column=employee_column,
            value=employee_access.id_card,
        )
        ws.cell(row=2, column=employee_column).fill = PatternFill(
            "solid", fgColor="ffa4ffa4"
        )
        for access_data in employee_access.access_data_list:
            if access_data.exit_out is None:
                logger.error(
                    f"No exit for enter event: {employee_access.name}, {employee_access.unit}, {access_data.enter_in}, {access_data.enter_in_turnstyle}, {access_data.restricted_area}"
                )
                continue
            e_date = date(
                access_data.exit_out.year,
                access_data.exit_out.month,
                access_data.exit_out.day,
            )
            row = date_row[e_date]
            value = ws.cell(row=row, column=employee_column).value
            access_seconds = (access_data.exit_out - access_data.enter_in).seconds
            if value is None:
                ws.cell(
                    row=row,
                    column=employee_column,
                    value=f"={access_seconds}",
                )
            else:
                ws.cell(
                    row=row,
                    column=employee_column,
                    value=f"{value}+{access_seconds}",
                )
    # Add SUM values to total cells
    for ws in wb.worksheets:
        total_row = ws.max_row + 1
        total_column = ws.max_column + 1
        if total_row <= 2 or total_column <= 2:
            ws.cell(row=1, column=1, value="total_seconds")
            ws.cell(row=3, column=1, value="hours")
            ws.cell(row=4, column=1, value="minutes")
            ws.cell(row=5, column=1, value="seconds")
            ws.cell(row=7, column=1, value="[HH]:MM:SS")
            ws.cell(row=1, column=2, value=123456)
            ws.cell(row=3, column=2, value="=INT(B1/60/60)")
            ws.cell(row=4, column=2, value="=INT((B1-B3*60*60)/60)")
            ws.cell(row=5, column=2, value="=B1-B3*60*60-B4*60")
            ws.cell(
                row=7,
                column=2,
                value='=TEXT(B3,"00")&":"&TEXT(B4,"00")&":"&TEXT(B5,"00")',
            )
            continue
        for row in range(3, total_row):
            ws.cell(
                row=row,
                column=total_column,
                value=f"=SUM({get_column_letter(2)}{row}:{get_column_letter(total_column - 1)}{row})",
            )
            ws.cell(row=row, column=total_column).fill = PatternFill(
                "solid", fgColor="00C0C0C0"
            )
        total_column += 1
        for column in range(2, total_column):
            ws.cell(
                row=total_row,
                column=column,
                value=f"=SUM({get_column_letter(column)}{3}:{get_column_letter(column)}{total_row - 1})",
            )
            ws.cell(row=total_row, column=column).fill = PatternFill(
                "solid", fgColor="00C0C0C0"
            )
        ws.freeze_panes = "B3"
    wb.save("result.xlsx")
    wb.close()


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        logger.error("No report filenames or folders provided.")
        exit(1)
    main(sys.argv[1:])
